from openpyxl import load_workbook
from pprint import pprint
from pymongo import MongoClient, TEXT

client = MongoClient('mongodb://broadcast:agestado@cluster0-shard-00-00-umdst.mongodb.net:27017,cluster0-shard-00-01-umdst.mongodb.net:27017,cluster0-shard-00-02-umdst.mongodb.net:27017/test?ssl=true&replicaSet=Cluster0-shard-0&authSource=admin&retryWrites=true')

db = client['db-broadcastprodutos']

coll1 = db['coll-derivativos']
wb1 = load_workbook('Base Cotacao_3.xlsx')
ws1 = wb1['Derivativos']

coll2 = db['coll-faq']
wb2 = load_workbook('Base_FAQ_1.xlsx')
ws2= wb2['FAQ']



def trataServRT(data):
    serv_rt=[]
    p = data.split('//')

    for i in p:
        
        p1 = i.split('=')

        if len(p1) > 1:
            plataforma = p1[0]

        aux = p1[1].split('/')

        for ii in aux:

            p2 = ii.split('--')

            if len(p2) == 4:
                serv={
                    "plataforma": plataforma,
                    "codserv": p2[0],
                    "descserv":p2[1],
                    "feeprof":p2[2],
                    "demo":p2[3]
                }
                serv_rt.append(serv)
            else:
                serv={
                    "plataforma": p1[0],
                    "codserv": p2[0],
                    "descserv":p2[1],
                    "feeprof":p2[2],
                    "feenprof": p2[3],
                    "demo":p2[4]
                }
                serv_rt.append(serv)
    return serv_rt

def trataServDelay(data):
    serv_delay=[]

    p = data.split('//')

    for i in p:
        
        p1 = i.split('=')

        if len(p1) > 1:
            plataforma = p1[0]
        
            p2 = p1[1].split('--')

            if len(p2) == 2:
                
                serv={
                    "plataforma": plataforma,
                    "servs":p2[0],
                    "delay":p2[1]
                }
                serv_delay.append(serv)
            
            else:
                serv={
                    "plataforma": plataforma,
                    "codserv": p2[0],
                    "descserv":p2[1],
                    "feeprof":p2[2],
                    "demo": p2[3],
                    "delay":p2[4]
                }
                serv_delay.append(serv)
        
        else:
            p2 = p1[0].split('--')

            serv={
                "servs":p2[0],
                "delay":p2[1]
            }
            serv_delay.append(serv)
               

    return serv_delay


#### Bloco de upload de planilha de FAQ ###
many_post=[]
row=2
while ws2.cell(row=row, column=1).value != None:

    post =  {
        "_id": ws2.cell(row=row,column=1).value,
        "pergunta":ws2.cell(row=row, column=2).value,
        "resposta":ws2.cell(row=row, column=3).value
    }

    many_post.append(post)
    row=row+1

coll2.insert_many(many_post)
coll2.create_index([("pergunta", TEXT),("resposta",TEXT)], name="faq_index")


#### Bloco de upload de planilha de DErivativos ###
many_post=[]
row=2
while ws1.cell(row=row, column=1).value != None:

    serv_rt = trataServRT(ws1.cell(row=row, column=10).value)
    serv_delay = trataServDelay(ws1.cell(row=row, column=11).value)

    post =  {
        "mercadoria": ws1.cell(row=row, column=7).value,
        "fonte":ws1.cell(row=row, column=1).value,
        "mercado":ws1.cell(row=row, column=6).value,
        "desc_papel":ws1.cell(row=row, column=3).value,
        "codbroad":ws1.cell(row=row, column=2).value,
        "codbolsa":ws1.cell(row=row, column=4).value,
        "tpinst":ws1.cell(row=row, column=5).value,
        "cod_pag":ws1.cell(row=row, column=9).value,
        "pag":ws1.cell(row=row, column=8).value,
        "serv_rt":serv_rt,
        "serv_delay":serv_delay,
    }

    many_post.append(post)
    row=row+1

coll1.insert_many(many_post)
coll1.create_index([("mercadoria", TEXT),("fonte",TEXT),("mercado", TEXT),("tpinst", TEXT),("desc_papel", TEXT),("codbroad", TEXT),("codbolsa", TEXT),("pag", TEXT)], name="derivativos_index")
#pprint(many_post)

#############

'''
post =  {
    "mercadoria": "Café",
    "fonte":"BMF",
    "mercado":"Futuro",
    "desc_papel":"Café Arábica Rolagem",
    "codbroad":"CR1",
    "codbolsa":"CR1",
    "tpinst":,
    "cod_pag":,
    "pag":,
    "serv_rt":[],
    "serv_delay":[],
    }
'''

## Serviços RT ##

# 1º Split por  '//' para separar as plaformas e seus respectivos serviços ||  p[0] = Broadcast e p[1] = Wallboard
# 2º Split por '=' para pegar a plataforma || p1[0] = Plataforma e p1[1]=Serviços
# 3º Split por '/' para separar os serviços   
# 4º Split por '--' para separar os campos dos serviços

    # Se a quantidade de posições for igual a 4 tomar como paramentro:
    # p2[0] = Cód Serv
    # p2[1] = Desc Serv
    # p2[2] = Fee
    # p2[3] = Dias Demo
'''
    serv_rt={
        "plataforma": p1[0],
        "codserv": p2[0],
        "descserv":p2[1],
        "fee":p2[2],
        "demo":p2[3]
    }
'''
    # Se a quantidade de posições for igual a 5 tomar como paramentro:
    # p2[0] = Cód Serv
    # p2[1] = Desc Serv
    # p2[2] = Fee Profissional
    # p2[3] = Fee Não Profissional
    # p2[4] = Dias Demo
'''
    serv_rt={
        "plataforma": p1[0],
        "codserv": p2[0],
        "descserv":p2[1],
        "feeprof":p2[2],
        "feenprof: p[3]
        "demo":p2[4]
    }
'''

## Serviços Delay ##

# 1º Split por  '//' para separar as plaformas e seus respectivos serviços ||  p[0] = Broadcast e p[1] = Wallboard
# 2º Split por '=' para pegar a plataforma || p1[0] = Plataforma e p1[1]=Serviços
# 3º Split por '/' para separar os serviços   
# 4º Split por '--' para separar os campos dos serviços

    # Se a quantidade de posições for igual a 2 tomar como paramentro:
    # p2[0] = Disp. todos pacotes
    # p2[1] = Tempo Delay

        # Se houver plataforma definida
'''
        serv_delay={
            "plataforma": p1[0],
            "servs":p2[0],
            "delay":p2[1]
        }
'''
        # Se Não houver plataforma definida
'''
        serv_delay={
            "servs":p2[0],
            "delay":p2[1]
        }
'''

    # Se a quantidade de posições for igual a 5 tomar como paramentro:
    # p2[0] = Cód Serv
    # p2[1] = Desc Serv
    # p2[2] = Fee
    # p2[3] = Dias Demo
'''
    serv_delay={
        "plaforma":p1[0],
        "codserv": p2[0],
        "descserv":p2[1],
        "fee":p2[2],
        "demo":p2[3],
        "delay":p2[4]
    }
'''