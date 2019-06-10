from openpyxl import load_workbook
from dbmongo import coll 
wb2 = load_workbook('D:\Mercados\Cotação\Base Cotacao.xlsx')

#print(wb2.sheetnames)

ws= wb2['Derivativos']

#print(ws['A3'].value)

#print(ws.cell(row=4, column=8).value)


#teste = ws.cell(row=67, column=8).value.split('/') # Separa os Serviços
#print(teste)
#print(len(teste))


row=2
col=1
many_post=[]


while ws.cell(row=row, column=1).value != None:
    rt=[]
    # Separa as Plataformas
    # servs_rt[0] == "plataforma":["BPRO", "BPRO WEB", "BAGRO", "BAGRO WEB", "TN"],
    # servs_rt[1] == "plataforma": WB
    servs_rt = ws.cell(row=row, column=10).value.split('//')
    print(servs_rt)   
    # Separa os Serviços por Plataformas
    if len(servs_rt) < 1:
        serv_bpro = servs_rt[0].split(';')
        serv_bw = servs_rt[1].split(';') 
        #print('serv_bpro: {}'.format(serv_bpro))

        if ws.cell(row=row, column=2).value == "BMF":
            
            for x in serv_bpro:
                serv=[]
                #Separa os Campos dos Serviços RT
                # Cód , DescServ, NProf, Prof, DiasDemo
                #print("x: {}".format(x))
                serv = x.split('*')
                print('serv: {}'.format(serv))
                doc={  
                    "plataforma":["BPRO", "BPRO WEB", "BAGRO", "BAGRO WEB", "TN"],
                    "cod":serv[0],
                    "desc":serv[1],
                    "fee":{
                        "nprof":serv[2],
                        "prof":serv[3]
                        },
                    "demo":serv[4]
                }
            
                rt.append(doc)
            
            for x in serv_bw:
                serv=[]
                #Separa os Campos dos Serviços RT
                # Cód , DescServ, fee, DiasDemo
                serv = x.split('*')

                doc={  
                    "plataforma":"WB",
                    "cod":serv[0],
                    "desc":serv[1],
                    "fee":serv[2],
                    "demo":serv[3]
                }
            
                rt.append(doc)

        if ws.cell(row=row, column=2).value != "BMF":
            for x in serv_bpro:
                serv=[]
                #Separa os Campos
                # Cód , DescServ, Fee, DiasDemo
                serv = x.split('*')

                serv_rt={  
                    "plataforma":["BPRO", "BPRO WEB", "BAGRO", "BAGRO WEB", "TN"],
                    "cod":serv[0],
                    "desc":serv[1],
                    "fee":serv[2],
                    "demo":serv[3]
                }
            
                rt.append(serv_rt)

            for x in serv_bw:
                serv=[]
                #Separa os Campos dos Serviços RT
                # Cód , DescServ, fee, DiasDemo
                serv = x.split('*')

                doc={  
                    "plataforma":"WB",
                    "cod":serv[0],
                    "desc":serv[1],
                    "fee":serv[2],
                    "demo":serv[3]
                }
            
                rt.append(doc)

    else:
        serv_bpro = servs_rt[0].split(';')
        #print('serv_bpro: {}'.format(serv_bpro))
        if ws.cell(row=row, column=2).value != "BMF":
            for x in serv_bpro:
                serv=[]
                #Separa os Campos
                # Cód , DescServ, Fee, DiasDemo
                serv = x.split('*')

                serv_rt={  
                    "plataforma":["BPRO", "BPRO WEB", "BAGRO", "BAGRO WEB", "TN"],
                    "cod":serv[0],
                    "desc":serv[1],
                    "fee":serv[2],
                    "demo":serv[3]
                }
            
                rt.append(serv_rt)

    post =  {
        "mercadoria": ws.cell(row=row, column=1).value,
        "fonte":ws.cell(row=row, column=2).value,
        "mercado":ws.cell(row=row, column=3).value,
        "desc_papel":ws.cell(row=row, column=4).value,
        "codbolsa":ws.cell(row=row, column=5).value,
        "codbroad":ws.cell(row=row, column=6).value,
        "pag_perm":ws.cell(row=row, column=10).value,
        "tp_instr":ws.cell(row=row, column=7).value,
        "serv":{
            "rt":rt,
            "delay":ws.cell(row=row, column=9).value,
            } 
        }

        #print(post)
    #coll.insert_one(post).insertd_id
    many_post.append(post)
    row=row+1

coll.insert_many(many_post)

'''
        if ws.cell(row=row, column=2).value == "BMF":
            
            for x in serv_bpro:
                serv=[]
                #Separa os Campos dos Serviços RT
                # Cód , DescServ, NProf, Prof, DiasDemo
                #print("x: {}".format(x))
                serv = x.split('*')
                print('serv: {}'.format(serv))
                doc={  
                    "plataforma":["BPRO", "BPRO WEB", "BAGRO", "BAGRO WEB", "TN"],
                    "cod":serv[0],
                    "desc":serv[1],
                    "fee":{
                        "nprof":serv[2],
                        "prof":serv[3]
                        },
                    "demo":serv[4]
                }
            
                rt.append(doc)
'''            
        

'''
post =  {
    "mercadoria": "Café",
    "fonte":"BMF",
    "mercado":"Futuro",
    "desc_papel":"Café Arábica Rolagem",
    "codbroad":"CR1",
    "codbolsa":"CR1",
    "tpinst":,
    "cod_pag":,
    "pag":,
    "serv_rt":[],
    "serv_delay":[],
    }
'''

## Serviços RT ##

# 1º Split por  '//' para separar as plaformas e seus respectivos serviços ||  p[0] = Broadcast e p[1] = Wallboard
# 2º Split por '=' para pegar a plataforma || p1[0] = Plataforma e p1[1]=Serviços
# 3º Split por '/' para separar os serviços   
# 4º Split por '--' para separar os campos dos serviços

     # Se a quantidade de posições for igual a 4 tomar como paramentro:
     # p2[0] = Cód Serv
     # p2[1] = Desc Serv
     # p2[2] = Fee
     # p2[3] = Dias Demo
    '''
    serv_rt={
        "plataforma": p1[0],
        "codserv": p2[0],
        "descserv":p2[1],
        "fee":p2[2],
        "demo":p2[3]
    }
    '''
     # Se a quantidade de posições for igual a 5 tomar como paramentro:
     # p2[0] = Cód Serv
     # p2[1] = Desc Serv
     # p2[2] = Fee Profissional
     # p2[3] = Fee Não Profissional
     # p2[4] = Dias Demo
    '''
    serv_rt={
        "plataforma": p1[0],
        "codserv": p2[0],
        "descserv":p2[1],
        "feeprof":p2[2],
        "feenprof: p[3]
        "demo":p2[4]
    }
    '''

## Serviços Delay ##

# 1º Split por  '//' para separar as plaformas e seus respectivos serviços ||  p[0] = Broadcast e p[1] = Wallboard
# 2º Split por '=' para pegar a plataforma || p1[0] = Plataforma e p1[1]=Serviços
# 3º Split por '/' para separar os serviços   
# 4º Split por '--' para separar os campos dos serviços

    # Se a quantidade de posições for igual a 2 tomar como paramentro:
    # p2[0] = Disp. todos pacotes
    # p2[1] = Tempo Delay

        # Se houver plataforma definida
        '''
        serv_delay={
            "plataforma": p1[0],
            "servs":p2[0],
            "delay":p2[1]
        }
        '''
        # Se Não houver plataforma definida
        '''
        serv_delay={
            "servs":p2[0],
            "delay":p2[1]
        }
        '''

    # Se a quantidade de posições for igual a 5 tomar como paramentro:
    # p2[0] = Cód Serv
    # p2[1] = Desc Serv
    # p2[2] = Fee
    # p2[3] = Dias Demo
    '''
    serv_delay={
        "plaforma":p1[0],
        "codserv": p2[0],
        "descserv":p2[1],
        "fee":p2[2],
        "demo":p2[3],
        "delay":p2[4]
    }
    '''
